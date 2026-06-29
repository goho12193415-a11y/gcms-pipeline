#!/usr/bin/env python3
"""Compare pipeline v2 (review-optimized) with user manual identification."""
import openpyxl, re

# ---- Load pipeline v2 results ----
PIPE = r"C:\Users\go ho\Desktop\gcms_pipeline_v2\output\GCMS_Results_20260627_233230.xlsx"
wb = openpyxl.load_workbook(PIPE, data_only=True)
ws = wb['Results']

# Build column index from header
cols = {}
for c in range(1, ws.max_column + 1):
    h = str(ws.cell(row=1, column=c).value or '')
    cols[h] = c

pipe = []
for r in range(2, ws.max_row + 1):
    sample = ws.cell(row=r, column=cols.get('Sample', 1)).value
    if not sample or 'xt6.26' not in str(sample):
        continue
    rank = ws.cell(row=r, column=cols.get('Rank', 999)).value
    # Only use Rank=1 (top match)
    if rank and rank != 1:
        continue
    rt = ws.cell(row=r, column=cols.get('RT_min', 3)).value
    name = ws.cell(row=r, column=cols.get('Compound_Name', 13)).value
    fmf = ws.cell(row=r, column=cols.get('FMF', 15)).value
    rmf = ws.cell(row=r, column=cols.get('RMF', 16)).value
    ri_d = ws.cell(row=r, column=cols.get('RI_Diff', 17)).value
    status = ws.cell(row=r, column=cols.get('Status', 11)).value
    if rt and name and name != 'Unknown':
        pipe.append({'rt': float(rt), 'name': str(name).strip(),
                     'fmf': fmf or 0, 'rmf': rmf or 0,
                     'ri_diff': ri_d, 'status': str(status)})

# ---- Load user manual results ----
USER = r"C:\Users\go ho\Desktop\桌面\xt_filled.xlsx"
uwb = openpyxl.load_workbook(USER, data_only=True)
uws = uwb['Sheet1']
user = []
for r in range(6, uws.max_row + 1):
    rt = uws.cell(row=r, column=1).value
    name = uws.cell(row=r, column=9).value or uws.cell(row=r, column=10).value or ''
    area_pct = uws.cell(row=r, column=5).value or 0
    if rt and str(name).strip():
        user.append({'rt': float(rt), 'name': str(name).strip(),
                     'area_pct': float(area_pct)})

# ---- Name matching ----
def norm(s):
    return re.sub(r'[^a-z0-9]', '', s.lower().strip())

SYNONYMS = {
    '2-pentylfuran': 'furan, 2-pentyl-',
    '2-ethyl-1-hexanol': '1-hexanol, 2-ethyl-',
    '2-(2-butoxyethoxy)ethanol': 'ethanol, 1-(2-butoxyethoxy)-',
    '3,5-di-tert-butylphenol': 'phenol, 3,5-bis(1,1-dimethylethyl)-',
    'beta-cyclocitral': 'photocitral b',
    '2,2,4-trimethyl-1,3-pentanediol monoisobutyrate':
        'propanoic acid, 2-methyl-, 3-hydroxy-2,2,4-trimethylpentyl ester',
    '(e)-2-octenal': '2-octenal, (e)-',
    'beta-ionone': 'trans-.beta.-ionone',
    '2,4-decadienal': '2,4-decadienal, (e,e)-',
    'dimethylsilanediol': 'silanediol, dimethyl-',
    '2,4,6-trimethylpyridine': 'pyridine, 2,4,6-trimethyl-',
    'butylated hydroxytoluene': '2,4,6-tris(1,1-dimethylethyl)-4-methylcyclohexa-2,5-dien-1-one',
    # IUPAC naming variants
    'butanoic acid': 'butanoic acid',
}

def normalize_alkane(name):
    """Convert between IUPAC and common alkane naming.
    '2,6,10-Trimethyldodecane' <-> 'Dodecane, 2,6,10-trimethyl-'
    Returns canonical form: 'dodecane-2,6,10-trimethyl'
    """
    n = name.lower().strip()
    # Pattern 1: "Dodecane, 2,6,10-trimethyl-"
    m = re.match(r'([a-z]+),?\s*([\d,]+-[a-z]+)', n)
    if m:
        base = m.group(1).strip().rstrip(',')
        sub = m.group(2).strip().rstrip('-')
        return f'{base}-{sub}'
    # Pattern 2: "2,6,10-Trimethyldodecane"
    m = re.match(r'([\d,]+-[a-z]+)([a-z]+)', n)
    if m:
        sub = m.group(1)
        base = m.group(2)
        return f'{base}-{sub}'
    return n

def names_match(u_name, p_name):
    u, p = u_name.lower().strip(), p_name.lower().strip()
    if u == p: return True
    # Synonyms
    for uk, pv in SYNONYMS.items():
        if (u == uk and p == pv) or (u == pv and p == uk): return True
    # E/Z prefix/suffix
    u_core = re.sub(r'^\((e|z|e,e|z,z)\)-', '', u, flags=re.IGNORECASE)
    p_core = re.sub(r',\s*\((e|z)\)$', '', p, flags=re.IGNORECASE)
    if u_core == p_core: return True
    # IUPAC <-> common alkane name normalization
    if normalize_alkane(u) == normalize_alkane(p): return True
    # Long name fallback
    if len(u) > 15 and len(p) > 15:
        u_n, p_n = norm(u), norm(p)
        if u_n == p_n: return True
        if len(u_n) > 12 and (u_n in p_n or p_n in u_n): return True
    return False

# ---- Match ----
agree, disagree, user_only = [], [], []
matched = set()

for up in user:
    best, best_dist = None, 0.5
    for i, pp in enumerate(pipe):
        d = abs(pp['rt'] - up['rt'])
        if d < best_dist:
            best_dist = d
            best = (i, pp)
    if best:
        i, pp = best
        if names_match(up['name'], pp['name']):
            agree.append({'u': up, 'p': pp, 'drt': round(best_dist, 3)})
        else:
            disagree.append({'u': up, 'p': pp, 'drt': round(best_dist, 3)})
        matched.add(i)
    else:
        user_only.append(up)

pipe_only = [pp for i, pp in enumerate(pipe) if i not in matched]

# ---- Count by status ----
green_agree = sum(1 for a in agree if a['p']['status'] == 'CONFIRMED')
yellow_agree = sum(1 for a in agree if a['p']['status'] == 'REVIEW')
gray_agree = sum(1 for a in agree if a['p']['status'] == 'LOW')
green_disagree = sum(1 for d in disagree if d['p']['status'] == 'CONFIRMED')
yellow_disagree = sum(1 for d in disagree if d['p']['status'] == 'REVIEW')

# ---- Report ----
print(f"Pipeline (Rank=1): {len(pipe)} top-1 matches on xt6.26")
print(f"User manual:       {len(user)} identifications")
print()
print(f"{'='*65}")
print(f"  VERIFICATION RESULT")
print(f"{'='*65}")
print(f"  Agreement:       {len(agree):3d}")
print(f"    GREEN (auto):  {green_agree}  — you don't need to check these")
print(f"    YELLOW (check):{yellow_agree}  — would catch in review")
print(f"    GRAY (low):    {gray_agree}  — might miss in review")
print(f"  Disagreement:    {len(disagree):3d}")
print(f"    Pipeline GREEN but wrong: {green_disagree}  — BAD: auto-confirmed wrong!")
print(f"    Pipeline YELLOW, you'd fix:{yellow_disagree}  — caught in review")
print(f"  Your peaks missed:{len(user_only):3d}")
print()
print(f"  EFFECTIVE ACCURACY (you review YELLOW):")
eff = green_agree + yellow_agree
print(f"    {eff}/{len(user)} = {eff*100/len(user):.0f}%")
print(f"    Green auto-confirmed: {green_agree} — automatically correct")
print(f"    Green false-positive: {green_disagree} — critical to check")
print(f"    You need to review:   ~{len(disagree)} peaks (YELLOW+some GRAY)")

# ---- Agreements ----
print(f"\n{'='*65}")
print(f"  AGREEMENTS ({len(agree)})")
print(f"{'='*65}")
for a in sorted(agree, key=lambda x: x['u']['rt']):
    pp = a['p']
    ri = f" dRI={pp['ri_diff']:.0f}" if pp.get('ri_diff') else ""
    print(f"  RT={a['u']['rt']:5.1f} [{pp['status']:<11s}] {a['u']['name']:<38s} "
          f"RMF={pp['rmf']:.0f}{ri}")

# ---- Disagreements ----
print(f"\n{'='*65}")
print(f"  DISAGREEMENTS ({len(disagree)})")
print(f"{'='*65}")
for d in sorted(disagree, key=lambda x: x['u']['rt']):
    pp = d['p']
    ri = f" dRI={pp['ri_diff']:.0f}" if pp.get('ri_diff') else ""
    print(f"  RT={d['u']['rt']:5.1f} [{pp['status']:<11s}] dRT={d['drt']:.3f}")
    print(f"    User: {d['u']['name']}")
    print(f"    Pipe: {pp['name']}  (RMF={pp['rmf']:.0f} FMF={pp['fmf']:.0f}){ri}")

# ---- User-only ----
if user_only:
    print(f"\n{'='*65}")
    print(f"  MISSED BY PIPELINE ({len(user_only)})")
    print(f"{'='*65}")
    for u in sorted(user_only, key=lambda x: x['rt']):
        print(f"  RT={u['rt']:5.1f}  {u['name']:<40s}  area={u['area_pct']:.1f}%")
