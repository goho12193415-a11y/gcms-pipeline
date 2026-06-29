#!/usr/bin/env python3
"""
eval/score.py — Coverage judge for the GC-MS pipeline.
======================================================
The single source of truth for every change in the remediation plan.

Compares the pipeline's top-5 candidates (dumped to
output/<sample>_candidates.json by pipeline.py) against the user's manual
identification (xt_filled.xlsx), and reports:
  - top-1 coverage  (pipeline's rank-1 == user answer)
  - top-5 coverage  (user answer anywhere in pipeline top-5)  <-- THE metric
  - the miss list    (user compounds NOT in pipeline top-5, with pipeline top-1)

Usage:
    python eval/score.py
    python eval/score.py --candidates output/xt6.26_candidates.json
    python eval/score.py --rt-tol 0.30
"""
import argparse
import json
import re
from pathlib import Path

import openpyxl

PROJECT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_CANDIDATES = PROJECT_DIR / "output" / "xt6.26_candidates.json"
DEFAULT_GROUND_TRUTH = r"C:\Users\go ho\Desktop\桌面\xt_filled.xlsx"

# ---------------------------------------------------------------------------
# Name matching  (lifted from compare.py — the validated logic)
# ---------------------------------------------------------------------------
SYNONYMS = {
    '2-pentylfuran': 'furan, 2-pentyl-',
    '2-ethyl-1-hexanol': '1-hexanol, 2-ethyl-',
    '2-(2-butoxyethoxy)ethanol': 'ethanol, 1-(2-butoxyethoxy)-',
    '3,5-di-tert-butylphenol': 'phenol, 3,5-bis(1,1-dimethylethyl)-',
    'beta-cyclocitral': 'photocitral b',
    '2,2,4-trimethyl-1,3-pentanediol monoisobutyrate':
        'propanoic acid, 2-methyl-, 3-hydroxy-2,2,4-trimethylpentyl ester',
    '(e)-2-octenal': '2-octenal, (e)-',
    'beta-ionone': 'trans-.beta.-ionone',
    '2,4-decadienal': '2,4-decadienal, (e,e)-',
    'dimethylsilanediol': 'silanediol, dimethyl-',
    '2,4,6-trimethylpyridine': 'pyridine, 2,4,6-trimethyl-',
    'butylated hydroxytoluene':
        '2,4,6-tris(1,1-dimethylethyl)-4-methylcyclohexa-2,5-dien-1-one',
    'beta-cyclocitral':
        '1-cyclohexene-1-carboxaldehyde, 2,6,6-trimethyl-',
    # one user name -> several valid NIST aliases
    'beta-ionone': ('trans-.beta.-ionone',
                    '3-buten-2-one, 4-(2,6,6-trimethyl-1-cyclohexen-1-yl)-'),
}


def _strip_stereo(s):
    """Remove stereochemistry descriptors so EI-indistinguishable isomers
    (E/Z, cis/trans, optical) compare equal — per the project's stated
    physical limit that the quadrupole cannot separate them."""
    s = s.lower().strip()
    # leading: (e)-, (z)-, (e,e)-, (2e,4e)-, trans-, cis-, (+)-, (-)-
    s = re.sub(r'^\(?\d*[ez](,\d*[ez])*\)?-', '', s)
    s = re.sub(r'^(trans|cis)-', '', s)
    s = re.sub(r'^\(?[+-]\)?-', '', s)
    # trailing: ", (e)-", ", (z)-", ", (e,e)-" (with or without final dash)
    s = re.sub(r',\s*\(?\d*[ez](,\d*[ez])*\)?-?$', '', s)
    return s.strip().strip(',').strip()


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', s.lower().strip())


def normalize_alkane(name):
    """Canonicalize IUPAC <-> common alkane naming.
    '2,6,10-Trimethyldodecane' <-> 'Dodecane, 2,6,10-trimethyl-' -> 'dodecane-2,6,10-trimethyl'
    """
    n = name.lower().strip()
    m = re.match(r'([a-z]+),?\s*([\d,]+-[a-z]+)', n)
    if m:
        base = m.group(1).strip().rstrip(',')
        sub = m.group(2).strip().rstrip('-')
        return f'{base}-{sub}'
    m = re.match(r'([\d,]+-[a-z]+)([a-z]+)', n)
    if m:
        return f'{m.group(2)}-{m.group(1)}'
    return n


def names_match(u_name, p_name):
    u, p = u_name.lower().strip(), p_name.lower().strip()
    if u == p:
        return True
    for uk, pv in SYNONYMS.items():
        aliases = pv if isinstance(pv, (list, tuple)) else (pv,)
        if (u == uk and p in aliases) or (p == uk and u in aliases):
            return True
    us, ps = _strip_stereo(u), _strip_stereo(p)
    if us == ps:
        return True
    if normalize_alkane(us) == normalize_alkane(ps):
        return True
    if normalize_alkane(u) == normalize_alkane(p):
        return True
    if len(u) > 15 and len(p) > 15:
        u_n, p_n = _norm(u), _norm(p)
        if u_n == p_n:
            return True
        if len(u_n) > 12 and (u_n in p_n or p_n in u_n):
            return True
    return False


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------
def load_ground_truth(path):
    """Load user manual IDs. English name = col 9, Apex RT = col 1, data from row 6."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    user = []
    for r in range(6, ws.max_row + 1):
        rt = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=9).value or ws.cell(row=r, column=10).value or ''
        area_pct = ws.cell(row=r, column=5).value or 0
        if rt and str(name).strip():
            user.append({'rt': float(rt), 'name': str(name).strip(),
                         'area_pct': float(area_pct or 0)})
    return user


def load_candidates(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------
def score(user, peaks, rt_tol=0.30, top_n=5):
    """For each user compound, find nearest pipeline peak by RT and check
    whether the user's answer appears at rank-1 and within top-N. The 'top5'
    key holds the within-top-N result (name kept for backward compatibility)."""
    results = []
    for up in user:
        best, best_d = None, rt_tol
        for pk in peaks:
            d = abs(pk['rt'] - up['rt'])
            if d < best_d:
                best_d, best = d, pk
        if best is None:
            results.append({'user': up, 'peak': None, 'top1': False,
                            'top5': False, 'drt': None})
            continue
        cands = best.get('candidates', [])
        top1 = bool(cands) and names_match(up['name'], cands[0]['name'])
        topn = any(names_match(up['name'], c['name']) for c in cands[:top_n])
        results.append({'user': up, 'peak': best, 'top1': top1,
                        'top5': topn, 'drt': round(best_d, 3)})
    return results


def main():
    ap = argparse.ArgumentParser(description='GC-MS coverage judge')
    ap.add_argument('--candidates', default=str(DEFAULT_CANDIDATES))
    ap.add_argument('--ground-truth', default=DEFAULT_GROUND_TRUTH)
    ap.add_argument('--rt-tol', type=float, default=0.30)
    ap.add_argument('--top-n', type=int, default=5, help='review window depth')
    args = ap.parse_args()

    user = load_ground_truth(args.ground_truth)
    peaks = load_candidates(args.candidates)
    res = score(user, peaks, rt_tol=args.rt_tol, top_n=args.top_n)

    n = len(res)
    n_top1 = sum(1 for r in res if r['top1'])
    n_top5 = sum(1 for r in res if r['top5'])
    n_nort = sum(1 for r in res if r['peak'] is None)

    print("=" * 66)
    print(f"  COVERAGE  (candidates: {Path(args.candidates).name}, rt_tol={args.rt_tol})")
    print("=" * 66)
    print(f"  User manual IDs:      {n}")
    print(f"  No peak within RT tol:{n_nort}")
    print(f"  top-1 coverage:       {n_top1}/{n} = {n_top1*100/n:.0f}%")
    print(f"  top-{args.top_n} coverage:       {n_top5}/{n} = {n_top5*100/n:.0f}%   <== METRIC")
    print()
    print("=" * 66)
    print(f"  MISSES  (user answer NOT in pipeline top-5)")
    print("=" * 66)
    for r in sorted((x for x in res if not x['top5']), key=lambda x: x['user']['rt']):
        up = r['user']
        if r['peak'] is None:
            print(f"  RT={up['rt']:5.2f}  {up['name']:<34s}  [no peak within {args.rt_tol}]")
            continue
        top = r['peak']['candidates'][0] if r['peak']['candidates'] else {}
        print(f"  RT={up['rt']:5.2f}  {up['name']:<34s}  dRT={r['drt']:.3f}")
        print(f"           pipe top-1: {top.get('name', '-'):<34s} "
              f"RMF={top.get('rmf', 0)} (area {up['area_pct']:.1f}%)")
    return n_top1, n_top5, n


if __name__ == "__main__":
    main()
