#!/usr/bin/env python3
"""
GC-MS Auto-Identification Pipeline v2.0
========================================
NIST-compatible spectral matching with forward/reverse match factors,
2-stage inverted-ion-index pre-search, and 3-tier contaminant classification.

Core matching logic mirrors Thermo Fisher Xcalibur/TraceFinder + NIST MS Search:
  1. Pre-search funnel: base-peak gate -> top-12 ion counter -> top-500 candidates
  2. NIST-weighted dot product: w = mz * sqrt(I / I_max)
  3. Forward MF: all shared peaks (pure spectrum vs library)
  4. Reverse MF: library peaks only (tolerates co-elution / background)
  5. Confidence: max(MF, RMF); co-elution flag when RMF - MF > 150

Usage:
    python gcms_pipeline.py sample.txt output.xlsx
    python gcms_pipeline.py --batch "data/*.txt" ./results/

Requirements: numpy, scipy, openpyxl
Author: GC-MS Pipeline Contributors
License: MIT
"""

import os, sys, re, json, argparse
from collections import defaultdict, Counter

import numpy as np
from scipy.signal import savgol_filter, find_peaks
from scipy.ndimage import minimum_filter1d, gaussian_filter1d
from scipy.integrate import simpson

# ---- Config ----
AMDIS_LIB_DIR = r"C:\NIST26-EI-DEMO\AMDIS32\LIB"

class Config:
    threshold = 700       # minimum SI (MF or RMF, whichever higher)
    use_msl = False
    mf_high = 850         # high confidence threshold
    mf_medium = 750       # medium confidence threshold
    coelution_delta = 150  # RMF - MF > this => flag co-elution
    min_shared_ions = 6    # minimum shared peaks for valid match
    pre_max_candidates = 500  # max candidates for full spectral matching
    pre_min_shared = 5     # minimum weighted shared ions to pass pre-search
    # RI filtering
    ri_tolerance = 50      # allowed RI deviation (±units) for polar columns
    ri_tolerance_nonpolar = 30  # allowed RI deviation for non-polar columns
    ri_weight = 0.15       # RI score contribution weight (SI * (1-ri_weight) + ri_score * ri_weight)
    ri_column = 'DB-WAX'  # default column for RI lookup
    ri_db_path = ''        # path to RI database JSON

PEAK_DISTANCE = 3        # min distance between peaks (scans); ~1.5s @ 0.5s/scan
PROMINENCE_FACTOR = 0.1   # prominence threshold relative to peak height

# ---- Background / contaminant markers ----
AIR_IONS = {28, 32, 40, 44, 18, 17}
BLEED_IONS = {73, 147, 207, 267, 281, 355}
PHTHALATE_IONS = {149, 167, 279}

# Ions to strip from observed spectrum before matching
# (siloxane, alkane fragments, phthalate — ubiquitous contaminants)
CONTAMINANT_IONS = {
    57, 71, 73, 85, 99, 113, 147, 149, 167, 207,
    221, 267, 279, 281, 295, 327, 341, 355, 429
}

CONTAMINANTS = {'2,4-Di-tert-butylphenol', 'Butylated hydroxytoluene',
                'Diethyl phthalate', 'Dibutyl phthalate'}

# Compounds unlikely to persist in aqueous HS-SPME
AQUEOUS_IMPOSSIBLE = {
    'Ethyl acetate', 'Hexyl acetate', 'Bornyl acetate',
    '1-Octen-3-ol, acetate', 'Octanoic acid, ethyl ester',
    'Diethyl phthalate', 'Dibutyl phthalate',
}

# Approximate RT ranges on DB-5 MS (50 min, 40->280°C @5°C/min)
# Used as sanity check — compounds outside range are rejected
RT_RANGES = {
    # Alcohols
    '1-Pentanol': (4, 10), '1-Penten-3-ol': (3, 8), '3-Methyl-1-butanol': (3, 9),
    '1-Hexanol': (6, 13), '2-Hexanol': (5, 12), '3-Hexen-1-ol': (7, 14),
    '1-Heptanol': (8, 16), '2-Heptanol': (7, 14),
    '1-Octanol': (10, 18), '2-Octanol': (9, 16), '1-Octen-3-ol': (8, 16),
    '2-Ethylhexanol': (8, 17), '1-Nonanol': (12, 20), '2-Nonanol': (10, 18),
    'Linalool': (10, 18), 'alpha-Terpineol': (12, 20),
    'Benzyl alcohol': (9, 17), 'Phenylethyl alcohol': (10, 19),
    # Aldehydes
    'Hexanal': (5, 12), '2-Hexenal': (7, 14), 'Heptanal': (7, 15),
    'Octanal': (9, 17), 'Nonanal': (11, 19), 'Decanal': (13, 22),
    'Benzaldehyde': (8, 16), 'Phenylacetaldehyde': (9, 18),
    # Ketones
    '2-Heptanone': (6, 14), '2-Octanone': (8, 16), '2-Nonanone': (10, 18),
    '2-Undecanone': (13, 22), 'Acetophenone': (9, 17),
    '6,10,14-Trimethylpentadecan-2-one': (28, 35),
    # Esters
    'Ethyl hexanoate': (8, 15), 'Ethyl octanoate': (10, 18),
    # Acids (underivatized on polar/wax columns are late; DB-5 tailing)
    'Acetic acid': (2, 8), 'Butanoic acid': (5, 12), 'Hexanoic acid': (8, 16),
    'Octanoic acid': (11, 20),
    # Terpenes
    'Limonene': (8, 15), 'alpha-Pinene': (6, 12), 'beta-Pinene': (7, 13),
    'beta-Caryophyllene': (18, 26), 'alpha-Humulene': (19, 27),
    # Furans / pyrans
    '2-Pentylfuran': (8, 15), '2-Furanmethanol': (6, 13),
    '5-Hydroxymethylfurfural': (20, 28),
    # Pyrazines
    '2-Methylpyrazine': (5, 12), '2,5-Dimethylpyrazine': (7, 14),
    '2,3,5-Trimethylpyrazine': (8, 16), '2-Ethyl-3,5-dimethylpyrazine': (10, 18),
    # Phenols
    'Phenol': (8, 15), 'Guaiacol': (10, 17), '4-Ethylguaiacol': (14, 22),
    'Eugenol': (15, 23), '4-Vinylguaiacol': (16, 24),
    # Sulfur
    'Dimethyl disulfide': (3, 8), 'Dimethyl trisulfide': (7, 14),
    'Methional': (7, 15),
    # Miscellaneous
    'Benzothiazole': (11, 19), 'Indole': (15, 24), 'Geosmin': (18, 26),
    '2-Methylisoborneol': (14, 22),
}


# ============================================================
# RETENTION INDEX (RI) CALCULATION
# ============================================================
# Default n-alkane RTs for DB-WAX (approximate, 40-280 C ramp)
# Calibrated from known compounds: 2-pentylfuran, benzaldehyde, beta-ionone
# User should replace with actual alkane standard RTs via --alkanes option
DEFAULT_ALKANE_RTS = {
    'DB-WAX': [
        (8.8, 1200), (15.4, 1400), (22.0, 1600), (28.6, 1800),
        (35.2, 2000), (41.8, 2200), (48.4, 2400), (55.0, 2600),
    ],
    'DB-5': [
        (7.0, 700), (8.5, 800), (10.0, 1000), (11.8, 1200),
        (13.8, 1400), (16.0, 1600), (18.5, 1800), (21.0, 2000),
        (23.5, 2200), (26.0, 2400), (28.5, 2600), (31.0, 2800),
    ],
}

def calc_ri_from_rt(rt, alkane_rts=None, column='DB-WAX'):
    """Convert retention time to Kovats Retention Index.

    Uses linear interpolation between consecutive n-alkanes:
        RI = 100 * [n + (RT_x - RT_n) / (RT_{n+1} - RT_n)]

    Args:
        rt: Retention time of unknown compound (min)
        alkane_rts: List of (RT, RI) pairs for n-alkanes, sorted by RT.
                    If None, uses DEFAULT_ALKANE_RTS for the column.
        column: Column type ('DB-WAX' or 'DB-5')
    Returns:
        Estimated Kovats RI, or None if RT is outside alkane range.
    """
    if alkane_rts is None:
        alkane_rts = DEFAULT_ALKANE_RTS.get(column, DEFAULT_ALKANE_RTS['DB-WAX'])

    if not alkane_rts or rt < alkane_rts[0][0] or rt > alkane_rts[-1][0]:
        # Extrapolate using first/last pair slope if possible
        if len(alkane_rts) >= 2:
            first, second = alkane_rts[0], alkane_rts[1]
            slope = (second[1] - first[1]) / (second[0] - first[0])
            if rt < first[0]:
                return round(first[1] + slope * (rt - first[0]))
            last, prev = alkane_rts[-1], alkane_rts[-2]
            slope = (last[1] - prev[1]) / (last[0] - prev[0])
            return round(last[1] + slope * (rt - last[0]))
        return None

    # Find bracketing alkane pair
    for i in range(len(alkane_rts) - 1):
        rt_n, ri_n = alkane_rts[i]
        rt_n1, ri_n1 = alkane_rts[i + 1]
        if rt_n <= rt <= rt_n1:
            n = ri_n // 100
            ri = 100 * n + 100 * (rt - rt_n) / (rt_n1 - rt_n)
            return round(ri)

    return None


def load_ri_database(db_path=None):
    """Load literature RI database from JSON file.

    Returns dict: {compound_name_lower: {column: ri_value}}
    """
    if db_path is None:
        here = os.path.dirname(os.path.realpath(__file__))
        db_path = os.path.join(here, 'ri_database.json')

    if not os.path.exists(db_path):
        print(f"[RI] Database not found: {db_path}")
        return {}

    with open(db_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    ri_db = {}
    for name, ri_values in data.get('compounds', {}).items():
        ri_db[name.lower()] = ri_values

    print(f"[RI] Loaded {len(ri_db)} compounds from {os.path.basename(db_path)}")
    return ri_db


def compute_ri_score(measured_ri, literature_ri, tolerance):
    """Score RI match: 1.0 for perfect match, decreasing with deviation.

    RI score = max(0, 1 - |RI_measured - RI_literature| / tolerance)
    """
    if measured_ri is None or literature_ri is None:
        return 0.5  # neutral if either is unknown
    delta = abs(measured_ri - literature_ri)
    if delta <= 5:
        return 1.0
    score = max(0.0, 1.0 - delta / tolerance)
    return score


def get_literature_ri(compound_name, ri_db, column='DB-WAX', cas=''):
    """Look up literature RI for a compound by CAS first, then by name."""
    # CAS lookup (most reliable)
    if cas and ri_db.get('by_cas'):
        entry = ri_db['by_cas'].get(cas)
        if entry and column in entry:
            return entry[column]
    # Name lookup (fallback)
    entry = ri_db.get(compound_name.lower())
    if entry and column in entry:
        return entry[column]
    # Try without column-specific key (AI estimate)
    if entry:
        ai_key = column + '_AI'
        if ai_key in entry:
            return entry[ai_key]
    return None


# ============================================================
# MSP LIBRARY PARSER
# ============================================================
def parse_msp(filepath):
    """Parse MSP (Mass Spectral Peak) format library.

    Example entry:
        Name: Hexanal
        Formula: C6H12O
        CASNO: 66-25-1
        Num Peaks: 18
        44 999; 56 450; 41 400; ...
    """
    compounds = []
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    raw_entries = []
    idx = 0
    while True:
        npos = content.find('Name:', idx)
        if npos == -1:
            break
        next_npos = content.find('\nName:', npos + 1)
        if next_npos == -1:
            raw_entries.append(content[npos:].strip())
            break
        raw_entries.append(content[npos:next_npos].strip())
        idx = next_npos + 1

    for entry in raw_entries:
        lines = entry.split('\n')
        name = ''
        formula = ''
        cas = ''
        peaks = []

        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith('Name: '):
                name = line[6:].strip()
            elif line.startswith('Formula: '):
                formula = line[9:].strip()
            elif line.startswith('CASNO: '):
                cas = line[7:].strip()
            elif line.startswith('Num Peaks:') or line.startswith('Synon:') or \
                 line.startswith('MW:') or line.startswith('RI:') or \
                 line.startswith('Comment:') or line.startswith('Contributor:'):
                continue
            else:
                numbers = re.findall(r'(\d+)\s+(\d+)', line)
                for mz_str, int_str in numbers:
                    peaks.append((int(mz_str), int(int_str)))

        if name and len(peaks) >= 5:
            top12 = [m for m, i in sorted(peaks, key=lambda x: -x[1])[:12]]
            fuzzy = set(top12)
            for m in top12:
                fuzzy.update({m - 1, m + 1})
            compounds.append({
                'name': name, 'formula': formula, 'cas': cas, 'peaks': peaks,
                'top8': top12, 'top8_fuzzy': fuzzy  # kept field name for compat
            })

    return compounds


# ============================================================
# MSL LIBRARY PARSER
# ============================================================
def parse_msl(filepath):
    """Parse AMDIS/NIST MSL format library."""
    compounds = []
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    entries = content.split('\nNAME: ')
    for i, entry in enumerate(entries):
        if i == 0 and entry.startswith('NAME: '):
            entry = entry[6:]
        if not entry.strip():
            continue

        lines = entry.strip().split('\n')
        name = lines[0].strip()
        formula = ''
        cas = ''
        peaks = []

        for line in lines[1:]:
            line = line.strip()
            if line.startswith('FORM:'):
                formula = line[5:].strip()
            elif line.startswith('CASNO:'):
                cas = line[6:].strip()
            else:
                m = re.findall(r'\(\s*(\d+)\s+(\d+)\s*\)', line)
                for mz_str, int_str in m:
                    peaks.append((int(mz_str), int(int_str)))

        if name and len(peaks) >= 5:
            top12 = [m for m, i in sorted(peaks, key=lambda x: -x[1])[:12]]
            fuzzy = set(top12)
            for m in top12:
                fuzzy.update({m - 1, m + 1})
            compounds.append({
                'name': name, 'formula': formula, 'cas': cas, 'peaks': peaks,
                'top8': top12, 'top8_fuzzy': fuzzy
            })

    return compounds


# ============================================================
# LIBRARY LOADER (with 2-stage index)
# ============================================================
def load_libraries(lib_dir=None):
    """Load spectral libraries and build 2-stage inverted index.

    Stage-1: base-peak index {bp_mz: [compound_indices]}
    Stage-2: top-12 ion mask per compound [(frozenset(top12), bp_mz), ...]

    Pre-search procedure:
      1. Gate by base peak (±1 Da): candidates from bp_index
      2. Weighted shared-ion scoring: low m/z (<80)=1pt, high m/z (>=80)=3pt
      3. Keep top-200 candidates with >=5 weighted score for full matching
    """
    all_compounds = []
    here = os.path.dirname(os.path.realpath(__file__))
    nist_path = os.path.join(here, 'nist_mainlib.msp')
    fallback_path = os.path.join(here, 'food_volatiles.msp')

    if os.path.exists(nist_path):
        comps = parse_msp(nist_path)
        all_compounds.extend(comps)
        print(f"[LIB] nist_mainlib.msp: {len(comps)} compounds (NIST 2014)")
    elif os.path.exists(fallback_path):
        comps = parse_msp(fallback_path)
        all_compounds.extend(comps)
        print(f"[LIB] food_volatiles.msp: {len(comps)} compounds")

    if lib_dir and os.path.isdir(lib_dir) and getattr(Config, 'use_msl', False):
        for fname in ['NISTFF.MSL', 'NISTEPA.MSL']:
            fp = os.path.join(lib_dir, fname)
            if os.path.isfile(fp):
                comps = parse_msl(fp)
                all_compounds.extend(comps)
                print(f"[LIB] {fname}: {len(comps)} compounds (MSL)")
    elif lib_dir and os.path.isdir(lib_dir):
        print("[LIB] MSL libraries available but not loaded (use --with-msl to enable)")

    # ---- Build 2-stage index ----
    bp_index = defaultdict(list)           # {base_peak_mz: [idx, ...]}
    ion_masks = []                          # [(frozenset(top12_ions), bp_mz), ...]

    for i, comp in enumerate(all_compounds):
        # Sort by intensity descending
        sorted_peaks = sorted(comp['peaks'], key=lambda x: -x[1])
        bp = int(round(sorted_peaks[0][0]))

        # Stage-1: base peak index (±1 Da tolerance)
        for d in (-1, 0, 1):
            bp_index[bp + d].append(i)

        # Stage-2: top-12 ion mask (±1 Da tolerance for matching)
        top12_set = set()
        for mz, _ in sorted_peaks[:12]:
            mz_i = int(round(mz))
            top12_set.update({mz_i - 1, mz_i, mz_i + 1})
        ion_masks.append((frozenset(top12_set), bp))

    total = len(all_compounds)
    print(f"[LIB] Total: {total} compounds, bp_index: {len(bp_index)} keys, "
          f"ion_masks: {len(ion_masks)} entries")

    # Load curated food volatiles library for fallback (Pass 3)
    food_path = os.path.join(here, 'food_volatiles.msp')
    if os.path.exists(food_path) and os.path.exists(nist_path):
        food_comps = parse_msp(food_path)
        Config.food_library = food_comps
        print(f"[LIB] Food library: {len(food_comps)} compounds (Pass 3 fallback)")
    else:
        Config.food_library = []

    return all_compounds, bp_index, ion_masks


# ============================================================
# NIST-COMPATIBLE SPECTRAL SIMILARITY (MF / RMF)
# ============================================================
def nist_similarity(observed, reference):
    """NIST MS Search compatible match factor calculation.

    Weighting:  w_i = mz_i * sqrt(I_i / I_max)
    Spectra are normalized to unit vectors for cosine similarity.

    Args:
        observed:  [(mz, intensity), ...]  — unknown spectrum
        reference: [(mz, intensity), ...]  — library spectrum

    Returns:
        (mf, rmf, n_shared, n_unknown, n_library)
          mf  — Forward Match Factor (0-999), all shared peaks
          rmf — Reverse Match Factor (0-999), library peaks only
    """
    # ---- Build weighted spectra ----
    def build_weighted(peaks):
        """Return {mz: weight} dict, NIST-weighted. Bins to 0.5 Da precision."""
        if not peaks:
            return {}
        max_i = max(i for _, i in peaks)
        if max_i <= 0:
            return {}
        result = {}
        for m, i in peaks:
            if i <= 0:
                continue
            # Bin to 0.5 Da (avoids collisions from coarse 1-Da rounding)
            mz = round(m * 2) / 2
            if mz < 26:
                continue
            w = mz * np.sqrt(i / max_i)
            if mz not in result or w > result[mz]:
                result[mz] = w
        return result

    # Remove low-m/z fragments
    obs_clean = [(m, i) for m, i in observed if int(round(m)) >= 26]

    obs_w = build_weighted(obs_clean)
    lib_w = build_weighted(reference)

    if len(obs_w) < 5 or len(lib_w) < 5:
        return 0, 0, len(obs_w), len(obs_w), len(lib_w)

    # ---- Find shared m/z (±0.5 Da tolerance, matches bin precision) ----
    lib_mz_set = set(lib_w.keys())
    shared = set()
    for mz in obs_w:
        for d in (-0.5, 0, 0.5):
            if mz + d in lib_mz_set:
                shared.add(mz)
                break

    n_shared = len(shared)
    if n_shared < 6:
        return 0, 0, n_shared, len(obs_w), len(lib_w)

    # ---- Forward MF: normalize over ALL peaks ----
    def normalize_to_unit(weights, restrict_to=None):
        """Normalize weight vector to unit length. Optionally restrict m/z set."""
        if restrict_to is not None:
            vals = [w for mz, w in weights.items() if mz in restrict_to]
        else:
            vals = list(weights.values())
        norm = np.sqrt(sum(v * v for v in vals))
        if norm < 1e-10:
            return {}
        return {mz: w / norm for mz, w in weights.items()
                if restrict_to is None or mz in restrict_to}

    obs_norm = normalize_to_unit(obs_w)
    lib_norm = normalize_to_unit(lib_w)

    mf = int(round(999 * sum(
        obs_norm.get(mz, 0) * lib_norm.get(mz, 0)
        for mz in shared
    )))
    mf = max(0, min(999, mf))

    # ---- Reverse MF: restrict to library m/z ----
    lib_peaks = set(lib_w.keys())
    obs_norm_rev = normalize_to_unit(obs_w, restrict_to=lib_peaks)
    lib_norm_rev = normalize_to_unit(lib_w)

    shared_rev = shared & lib_peaks
    if not shared_rev or not obs_norm_rev or not lib_norm_rev:
        rmf = 0
    else:
        rmf = int(round(999 * sum(
            obs_norm_rev.get(mz, 0) * lib_norm_rev.get(mz, 0)
            for mz in shared_rev
        )))
        rmf = max(0, min(999, rmf))

    return mf, rmf, n_shared, len(obs_w), len(lib_w)


# ============================================================
# PRE-SEARCH: 2-stage funnel
# ============================================================
# Background ions that should not be used as base peak for gating
BACKGROUND_BP_IONS = AIR_IONS | BLEED_IONS | {149, 167, 279}

def pre_search_candidates(clean_spectrum, bp_index, ion_masks, library):
    """2-stage funnel pre-search with background-aware base peak gating.

    Stage 1: base-peak gate (background-aware)
        If the observed base peak is an air/bleed/plasticizer ion,
        use up to 3 alternative base peaks from the first non-background
        ions. This prevents background-dominated spectra from failing
        the gate when the true compound's base peak is a different ion.

    Stage 2: top-12 ion counter
        Count how many of unknown's top-12 ions (±1 Da) match each candidate's
        top-12 ions.
    """
    if len(clean_spectrum) < 3:
        return []

    # Sorted by intensity
    sorted_obs = sorted(clean_spectrum, key=lambda x: -x[1])

    # Collect up to 3 effective base peaks, skipping background ions
    effective_bps = []
    for mz, _ in sorted_obs[:15]:
        mz_i = int(round(mz))
        if mz_i not in BACKGROUND_BP_IONS:
            if mz_i not in effective_bps:
                effective_bps.append(mz_i)
            if len(effective_bps) >= 3:
                break
    if not effective_bps:
        effective_bps = [int(round(sorted_obs[0][0]))]

    # Observed top-12 ion set (±1 Da tolerance)
    obs_ion_set = set()
    for mz, _ in sorted_obs[:12]:
        mz_i = int(round(mz))
        obs_ion_set.update({mz_i - 1, mz_i, mz_i + 1})

    # Stage 1: collect candidates from ALL effective base peaks
    candidates = set()
    for bp in effective_bps:
        for d in (-1, 0, 1):
            candidates.update(bp_index.get(bp + d, []))

    if not candidates:
        return []

    # Stage 2: count shared top-12 ions
    scored = []
    for idx in candidates:
        ion_set, _ = ion_masks[idx]
        shared = len(ion_set & obs_ion_set)
        if shared >= Config.pre_min_shared:
            scored.append((shared, idx))

    if not scored:
        return []

    scored.sort(reverse=True)
    return [idx for _, idx in scored[:Config.pre_max_candidates]]


# ============================================================
# DATA PARSER (Thermo Xcalibur TXT)
# ============================================================
def parse_thermo_txt(filepath):
    """Parse Thermo Xcalibur GC-MS text export."""
    rts, tics, spectra = [], [], []
    rt, tic, pkts, active = None, None, [], False

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            if s.startswith('ScanHeader'):
                if active and rt is not None:
                    rts.append(rt)
                    tics.append(tic or 0)
                    spectra.append(pkts)
                rt, tic, pkts, active = None, None, [], True
                continue
            if not active:
                continue
            if rt is None:
                m = re.search(r'start_time = ([\d.]+)', s)
                if m:
                    rt = float(m.group(1))
            m = re.search(r'integ_intens = ([\d.]+)', s)
            if m:
                tic = float(m.group(1))
            m = re.search(
                r'Packet # \d+, intensity = ([\d.eE+-]+), mass/position = ([\d.]+)',
                s
            )
            if m:
                pkts.append((float(m.group(2)), float(m.group(1))))
        if active and rt is not None:
            rts.append(rt)
            tics.append(tic or 0)
            spectra.append(pkts)

    return np.array(rts), np.array(tics), spectra


# ============================================================
# PEAK DETECTION
# ============================================================
def detect_peaks(rts, tics, spectra):
    """Detect peaks, filter background, integrate areas, extract spectra."""
    if len(tics) < 31:
        return []

    window = min(31, len(tics) // 2 * 2 + 1)
    if window < 5:
        return []

    tics_s = savgol_filter(tics, window, 2)
    baseline_window = min(401, len(tics) // 3 * 2 + 1)
    bl = minimum_filter1d(tics_s, baseline_window)
    bl = gaussian_filter1d(bl.astype(float), min(80, len(bl) // 4))
    corr = tics_s - bl
    corr[corr < 0] = 0

    if len(corr[corr > 0]) == 0:
        return []

    thresh = np.percentile(corr[corr > 0], 50)
    idxs, _ = find_peaks(corr, height=thresh, distance=PEAK_DISTANCE,
                          prominence=thresh * PROMINENCE_FACTOR)
    min_tic = np.median(tics) * 0.75
    idxs = [i for i in idxs if tics[i] > min_tic]

    results = []
    for idx in idxs:
        left = idx
        while left > 0 and corr[left] > corr[left - 1]:
            left -= 1
        right = idx
        while right < len(corr) - 1 and corr[right] > corr[right + 1]:
            right += 1

        x = rts[left:right + 1]
        y = tics[left:right + 1]
        base = np.linspace(tics[left], tics[right], len(y))
        yc = y - base
        yc[yc < 0] = 0
        area = float(simpson(yc, x)) if len(x) >= 3 else 0

        raw = spectra[idx]
        top = sorted(raw, key=lambda x: -x[1])[:40]

        results.append({
            'rt': round(rts[idx], 4),
            'tic': round(tics[idx], 0),
            'area': round(area, 0),
            'spectrum': top,
            'scan': int(idx),
        })

    return results


# ============================================================
# 3-TIER BACKGROUND CLASSIFICATION
# ============================================================
def classify_peak(mz_int_pairs):
    """3-tier classification of GC-MS peaks.

    L1: Natural products (aldehydes, ketones, alcohols, esters, terpenes...)
    L2: Suspected contaminants (branched alkanes, siloxanes, PEG, phthalates)
    L3: Confirmed background (air/water, column bleed) — discard

    Returns (tier, label, keep_for_matching)
    """
    mz_list = [int(round(m)) for m, i in mz_int_pairs[:15]]
    bp = mz_list[0] if mz_list else 0

    # == L3: Confirmed background (ignore entirely) ==
    air_count = sum(1 for m in AIR_IONS if m in mz_list[:8])
    if air_count >= 5:
        return ('L3', 'Air/Water', False)

    bleed_count = sum(1 for m in BLEED_IONS if m in mz_list)
    if bleed_count >= 2:
        return ('L3', 'Column Bleed', False)

    # == L2: Suspected contaminants (keep but flag) ==
    if 149 in mz_list[:5]:
        return ('L2', 'Plasticizer', True)

    if bp == 57 and all(m in mz_list for m in [43, 71, 85]):
        return ('L2', 'Branched Alkane', True)

    if bp == 73:
        return ('L2', 'Siloxane Derivative', True)

    if bp == 45 and len(mz_list) < 12:
        return ('L2', 'Glycol/PEG Derivative', True)

    # == L1: Natural products ==
    return ('L1', 'Natural Product', True)


# ============================================================
# MAIN PROCESSING
# ============================================================
def process_sample(filepath, library, bp_index, ion_masks, output_path,
                   ri_db=None, alkane_rts=None, ri_filter=False):
    """Run full pipeline on one GC-MS data file.

    Args:
        ri_db: Literature RI database {name: {column: ri_value}}
        alkane_rts: Alkane standard RTs [(rt, ri), ...]
        ri_filter: If True, filter/re-rank by RI deviation
    """
    name = os.path.basename(filepath)
    ri_column = Config.ri_column
    ri_tolerance = Config.ri_tolerance_nonpolar if ri_column == 'DB-5' else Config.ri_tolerance

    print(f"\n[Processing] {name}")
    if ri_filter:
        print(f"  RI filtering: {ri_column} ±{ri_tolerance}")

    # Parse
    rts, tics, spectra = parse_thermo_txt(filepath)
    print(f"  Scans: {len(rts)}, RT: {rts[0]:.1f}-{rts[-1]:.1f} min")

    # Detect & integrate
    peaks = detect_peaks(rts, tics, spectra)
    print(f"  Raw peaks: {len(peaks)}")

    # Identify
    results = []
    n_pre_searched = 0
    n_full_matched = 0

    for p in peaks:
        raw = p['spectrum']
        # Build clean spectrum
        clean = [(int(round(m)), int(i)) for m, i in raw
                 if int(round(m)) > 25 and i > 100]
        if len(clean) < 5:
            continue

        # 3-tier classification
        tier, peak_label, keep_for_matching = classify_peak(clean[:15])
        if not keep_for_matching:
            continue  # L3: skip entirely

        # Remove air ions from L1 for cleaner matching
        if tier == 'L1':
            air_count = sum(1 for m, i in clean[:12] if m in AIR_IONS)
            if air_count >= 6:
                continue

        # ---- 2-stage pre-search (2-pass) ----
        candidates = pre_search_candidates(clean, bp_index, ion_masks, library)
        n_pre_searched += 1

        if not candidates:
            continue

        # Pass 1: top 200 candidates (fast)
        pass1 = candidates[:200]
        # Pass 2: remaining 300 (fallback if no match in pass 1)
        pass2 = candidates[200:]

        # ---- Full NIST spectral matching ----
        matches = []
        tier_threshold = Config.mf_medium if tier == 'L1' else max(Config.threshold, Config.mf_high)

        for idx in pass1:
            comp = library[idx]
            mf, rmf, n_shared, n_unk, n_lib = nist_similarity(clean, comp['peaks'])
            si = max(mf, rmf)
            if si >= tier_threshold:
                matches.append((mf, rmf, si, n_shared, comp))

        # Pass 2: expanded search (only if pass 1 found nothing and pass 2 exists)
        if not matches and pass2:
            for idx in pass2:
                comp = library[idx]
                mf, rmf, n_shared, n_unk, n_lib = nist_similarity(clean, comp['peaks'])
                si = max(mf, rmf)
                if si >= tier_threshold:
                    matches.append((mf, rmf, si, n_shared, comp))

        # Pass 3: brute-force scan of curated food library (no BP gate)
        # For compounds whose library base peak is weak/absent in observed spectrum
        if not matches and hasattr(Config, 'food_library') and Config.food_library:
            obs_set = set()
            sorted_clean = sorted(clean, key=lambda x: -x[1])
            for mz, _ in sorted_clean[:12]:
                mz_i = int(round(mz))
                obs_set.update({mz_i - 1, mz_i, mz_i + 1})
            for fi, comp in enumerate(Config.food_library):
                fi_set = set()
                fi_sorted = sorted(comp['peaks'], key=lambda x: -x[1])
                for mz, _ in fi_sorted[:8]:
                    mz_i = int(round(mz))
                    fi_set.update({mz_i - 1, mz_i, mz_i + 1})
                if len(obs_set & fi_set) >= 5:
                    mf, rmf, n_shared, n_unk, n_lib = nist_similarity(clean, comp['peaks'])
                    si = max(mf, rmf)
                    if si >= tier_threshold:
                        matches.append((mf, rmf, si, n_shared, comp))

        if not matches:
            continue

        n_full_matched += 1

        # Filter aqueous-impossible (unless very high confidence)
        filtered = []
        for mf, rmf, si, n_shared, comp in matches:
            if comp['name'] in AQUEOUS_IMPOSSIBLE and max(mf, rmf) < 950:
                continue
            filtered.append((mf, rmf, si, n_shared, comp))

        if not filtered:
            continue

        # Exclude confirmed phthalate patterns
        has_phthalate_149 = any(int(round(m)) == 149 for m, i in clean[:20])
        has_phthalate_57 = any(int(round(m)) == 57 for m, i in clean[:20])
        if has_phthalate_149 and has_phthalate_57:
            continue

        # Sort by SI descending
        filtered.sort(key=lambda x: -x[2])

        # ---- RI re-ranking (if enabled) ----
        if ri_filter and ri_db:
            # Calculate measured RI for this peak
            ri_measured = calc_ri_from_rt(p['rt'], alkane_rts, ri_column)

            # Re-rank top matches with combined SI + RI score
            reranked = []
            for mf_i, rmf_i, si_i, n_sh_i, comp_i in filtered:
                ri_lit = get_literature_ri(comp_i['name'], ri_db, ri_column, cas=comp_i.get('cas', ''))
                ri_score = compute_ri_score(ri_measured, ri_lit, ri_tolerance)
                combined_si = int(round(si_i * (1 - Config.ri_weight) + ri_score * 999 * Config.ri_weight))

                if ri_filter and ri_lit is not None:
                    # Hard filter: reject if RI deviation > 2x tolerance
                    if ri_measured is not None and abs(ri_measured - ri_lit) > 2 * ri_tolerance:
                        continue

                reranked.append((combined_si, mf_i, rmf_i, si_i, n_sh_i, comp_i, ri_measured, ri_lit, ri_score))

            if reranked:
                reranked.sort(key=lambda x: -x[0])
                best = reranked[0]
                combined_si, mf, rmf, si, n_shared, comp, ri_measured, ri_lit, ri_score = best
            else:
                # All RI-filtered out; fall back to top SI match without RI
                mf, rmf, si, n_shared, comp = filtered[0]
                ri_measured = calc_ri_from_rt(p['rt'], alkane_rts, ri_column)
                ri_lit = get_literature_ri(comp['name'], ri_db, ri_column, cas=comp.get('cas', ''))
                ri_score = compute_ri_score(ri_measured, ri_lit, ri_tolerance)
                combined_si = si
        else:
            mf, rmf, si, n_shared, comp = filtered[0]
            combined_si = si
            ri_measured = None
            ri_lit = None
            ri_score = None

        # Exclude known contaminants
        if comp['name'] in CONTAMINANTS:
            continue

        # RT sanity check (skip if RI is available)
        if not ri_filter:
            rt_range = RT_RANGES.get(comp['name'])
            if rt_range and (p['rt'] < rt_range[0] or p['rt'] > rt_range[1]):
                continue

        # Co-elution flag
        coelution = (rmf - mf) > Config.coelution_delta

        # Confidence level (use combined_si when RI is active)
        eff_si = combined_si if ri_filter else si
        if eff_si >= Config.mf_high:
            confidence = 'High'
        elif eff_si >= Config.mf_medium:
            confidence = 'Medium'
        else:
            confidence = 'Low'

        results.append({
            'rt': p['rt'],
            'area': p['area'],
            'tic': p['tic'],
            'compound': comp['name'],
            'cas': comp.get('cas', ''),
            'formula': comp.get('formula', ''),
            'mf': mf,
            'rmf': rmf,
            'si': si,
            'n_shared': n_shared,
            'ri_measured': ri_measured,
            'ri_literature': ri_lit,
            'ri_score': round(ri_score, 2) if ri_score is not None else None,
            'combined_si': combined_si if ri_filter else None,
            'confidence': confidence,
            'coelution': coelution,
            'tier': tier,
            'peak_type': peak_label,
        })

    # ---- Deduplicate ----
    results.sort(key=lambda x: -x['si'])
    seen = set()
    final = []
    for r in results:
        key = (round(r['rt'], 1), r['compound'])
        if key not in seen:
            seen.add(key)
            final.append(r)
    final.sort(key=lambda x: x['rt'])

    # ---- Report ----
    _write_excel(final, output_path, name)

    n_total = len(results)
    n_final = len(final)
    n_high = sum(1 for r in final if r['confidence'] == 'High')
    n_med = sum(1 for r in final if r['confidence'] == 'Medium')
    n_coel = sum(1 for r in final if r.get('coelution'))
    print(f"  Pre-search: {n_pre_searched} peaks, full-matched: {n_full_matched}")
    print(f"  Identified: {n_final} (High:{n_high} Med:{n_med}"
          f" Co-elution:{n_coel})")
    return final


# ============================================================
# EXCEL OUTPUT
# ============================================================
def _write_excel(results, path, sample_name):
    """Generate formatted Excel report with MF/RMF columns."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sample_name[:31]

    hf = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    df = Font(name='Arial', size=10)
    da = Alignment(horizontal='center', vertical='center')
    dw = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bd = Border(left=Side('thin'), right=Side('thin'),
                top=Side('thin'), bottom=Side('thin'))

    # Color fills
    gf = PatternFill('solid', fgColor='C6EFCE')   # green: High
    yf = PatternFill('solid', fgColor='FFEB9C')    # yellow: Medium
    l2_fill = PatternFill('solid', fgColor='FFC000')  # orange: L2 contaminant
    l3_fill = PatternFill('solid', fgColor='FF9999')  # red: L3 background
    co_fill = PatternFill('solid', fgColor='D9B3FF')  # purple: co-elution

    has_ri = any(r.get('ri_measured') is not None for r in results)
    hdrs = ['No.', 'RT (min)', 'Compound', 'CAS', 'MF', 'RMF', 'SI',
            'N Shared', 'RI Meas.', 'RI Lit.', 'RI Score',
            'Confidence', 'Co-elution', 'Tier', 'Peak Type',
            'Peak Area', 'Peak Height', 'Formula']
    widths = [5, 9, 35, 15, 6, 6, 6, 8, 9, 9, 8,
              11, 10, 5, 16, 15, 15, 18]

    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = dw
        cell.border = bd
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    for i, r in enumerate(results, 2):
        vals = [
            i - 1, r['rt'], r['compound'], r['cas'],
            r.get('mf', ''), r.get('rmf', ''), r['si'],
            r.get('n_shared', ''),
            r.get('ri_measured', '') if r.get('ri_measured') is not None else '',
            r.get('ri_literature', '') if r.get('ri_literature') is not None else '',
            round(r['ri_score'], 2) if r.get('ri_score') is not None else '',
            r['confidence'], 'Y' if r.get('coelution') else '',
            r.get('tier', ''), r.get('peak_type', ''),
            r['area'], r['tic'], r.get('formula', '')
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = df
            cell.alignment = da
            cell.border = bd

        # Color coding
        tier = r.get('tier', 'L1')
        coel = r.get('coelution')

        if tier == 'L3':
            fill = l3_fill
        elif tier == 'L2':
            fill = l2_fill
        elif coel:
            fill = co_fill
        elif r['confidence'] == 'High':
            fill = gf
        elif r['confidence'] == 'Medium':
            fill = yf
        else:
            fill = None

        if fill:
            for c in range(1, len(hdrs) + 1):
                ws.cell(row=i, column=c).fill = fill

    n_cols = len(hdrs)
    ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(n_cols)}{len(results) + 1}'
    ws.freeze_panes = 'A2'
    wb.save(path)


# ============================================================
# CLI
# ============================================================
def main():
    ap = argparse.ArgumentParser(description='GC-MS Auto-Identification Pipeline v2.0')
    ap.add_argument('input', help='Input TXT file or glob pattern with --batch')
    ap.add_argument('output', help='Output Excel path or directory with --batch')
    ap.add_argument('--lib', default=AMDIS_LIB_DIR, help='AMDIS LIB directory path')
    ap.add_argument('--batch', action='store_true', help='Batch mode')
    ap.add_argument('--with-msl', action='store_true', help='Load AMDIS MSL libraries')
    ap.add_argument('--threshold', type=int, default=700,
                    help='Minimum SI threshold MF or RMF (default: 700)')
    ap.add_argument('--mf-high', type=int, default=850,
                    help='High confidence threshold (default: 850)')
    ap.add_argument('--mf-medium', type=int, default=750,
                    help='Medium confidence threshold (default: 750)')
    # RI options
    ap.add_argument('--ri-column', default='DB-WAX', choices=['DB-5', 'DB-WAX'],
                    help='Column type for RI lookup (default: DB-WAX)')
    ap.add_argument('--ri-db', default='', help='Path to RI database JSON')
    ap.add_argument('--ri-tolerance', type=int, default=50,
                    help='RI deviation tolerance (default: 50)')
    ap.add_argument('--with-ri', action='store_true',
                    help='Enable RI-based filtering during matching')
    ap.add_argument('--ri-no-filter', action='store_true',
                    help='Annotate RI but do not filter by it')
    ap.add_argument('--alkanes', default='',
                    help='Alkane RTs: C8_RT,C10_RT,C12_RT,... (comma-separated)')
    args = ap.parse_args()

    Config.threshold = args.threshold
    Config.mf_high = args.mf_high
    Config.mf_medium = args.mf_medium
    Config.use_msl = args.with_msl
    Config.ri_column = args.ri_column
    Config.ri_tolerance_nonpolar = 30 if args.ri_column == 'DB-5' else args.ri_tolerance
    Config.ri_tolerance = args.ri_tolerance
    if args.ri_db:
        Config.ri_db_path = args.ri_db

    # Load RI database
    ri_db = {}
    alkane_rts = None
    if args.with_ri:
        ri_db = load_ri_database(Config.ri_db_path if Config.ri_db_path else None)
        # Parse user-provided alkane RTs
        if args.alkanes:
            try:
                vals = [float(x.strip()) for x in args.alkanes.split(',')]
                carbons = list(range(len(vals) * 2 + 6, 5, -2))  # rough estimate
                alkane_rts = [(vals[i], 100 * (8 + i)) for i in range(len(vals))]
            except ValueError:
                print("[RI] Invalid alkane RT format, using defaults")
        if alkane_rts is None:
            alkane_rts = DEFAULT_ALKANE_RTS.get(args.ri_column)
        print(f"[RI] Column: {args.ri_column}, tolerance: ±{args.ri_tolerance}")
        print(f"[RI] Alkane RTs: {len(alkane_rts) if alkane_rts else 0} points")

    print("=" * 60)
    print("GC-MS Auto-Identification Pipeline v2.0")
    ri_status = f" + RI({args.ri_column})" if args.with_ri else ""
    print(f"  NIST MF/RMF matching{ri_status} | threshold={Config.threshold}"
          f" | MF high={Config.mf_high} med={Config.mf_medium}")
    print("=" * 60)

    library, bp_index, ion_masks = load_libraries(args.lib)

    if args.batch:
        import glob
        files = glob.glob(args.input)
        os.makedirs(args.output, exist_ok=True)
        for f in sorted(files):
            base = os.path.splitext(os.path.basename(f))[0]
            out = os.path.join(args.output, f'{base}_identified.xlsx')
            process_sample(f, library, bp_index, ion_masks, out,
                          ri_db=ri_db, alkane_rts=alkane_rts,
                          ri_filter=args.with_ri and not args.ri_no_filter)
    else:
        process_sample(args.input, library, bp_index, ion_masks, args.output,
                      ri_db=ri_db, alkane_rts=alkane_rts,
                      ri_filter=args.with_ri and not args.ri_no_filter)

    print("\nDone.")


if __name__ == '__main__':
    main()
