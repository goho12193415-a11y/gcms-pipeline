"""
Step 9: Results Export v2 — Review-optimized Excel
====================================================
  - Top-3 candidates per peak (not just top-1)
  - Auto-confirm rules (RMF>900 + RI<20 + common food volatile)
  - Contaminant auto-tagging (siloxanes, phthalates, column bleed)
  - Area descending sort (biggest peaks first)
  - Color-coded: GREEN=auto-confirmed, YELLOW=review needed,
                 RED=contaminant, GRAY=low confidence
"""
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import re

try:
    from config import TOP_N_CANDIDATES
except ImportError:
    TOP_N_CANDIDATES = 5


# ---- RI QC: does the top-1 ID's literature RI sit on this column's transform? ----
_WAX = None
_WAX_NORM = None


def _ri_canon(name):
    n = name.lower().strip()
    n = re.sub(r'^\(?\d*[ez](,\d*[ez])*\)?-', '', n)
    n = re.sub(r'^(trans|cis)-', '', n)
    n = re.sub(r',\s*\(?\d*[ez](,\d*[ez])*\)?-?$', '', n)
    return re.sub(r'[^a-z0-9]', '', n)


def _load_wax():
    global _WAX, _WAX_NORM
    if _WAX is not None:
        return
    import json
    from config import PROJECT_DIR
    try:
        d = json.load(open(PROJECT_DIR / "library" / "ri_dual_column.json", encoding='utf-8'))
        _WAX = {k.lower(): v for k, v in d.get('wax', {}).items()}
    except Exception:
        _WAX = {}
    _WAX_NORM = {}
    for k, v in _WAX.items():
        _WAX_NORM.setdefault(_ri_canon(k), v)


def _lit_ri(name):
    if not name:
        return None
    _load_wax()
    return _WAX.get(name.lower().strip()) or _WAX_NORM.get(_ri_canon(name))


_PRED = None
_PRED_NORM = None


def _load_pred():
    """CNN-predicted DB-WAX RI (library/ri_nist_full.json) — last-resort only."""
    global _PRED, _PRED_NORM
    if _PRED is not None:
        return
    import json
    from config import PROJECT_DIR
    try:
        d = json.load(open(PROJECT_DIR / "library" / "ri_nist_full.json", encoding='utf-8'))
        _PRED = {k.lower(): v for k, v in d.get('name_to_ri', {}).items()}
    except Exception:
        _PRED = {}
    _PRED_NORM = {}
    for k, v in _PRED.items():
        _PRED_NORM.setdefault(_ri_canon(k), v)


def _pred_ri(name):
    if not name:
        return None
    _load_pred()
    return _PRED.get(name.lower().strip()) or _PRED_NORM.get(_ri_canon(name))


def _fit_ri_transform(integrated_peaks, identification_results, rmf_min=850):
    """Self-calibrating per-sample: fit literatureRI = a*measuredRI + b from this
    sample's own confident top-1 hits (strong MS + has literature RI), so RI_Check
    works on ANY column/instrument without hard-coded constants. Returns (a, b)
    or None if too few consistent anchors to calibrate."""
    import numpy as np
    pts = []
    for i, p in enumerate(integrated_peaks):
        m = identification_results[i] if i < len(identification_results) else []
        if not m:
            continue
        mri = p.get('ri_measured'); lit = _lit_ri(m[0].get('name', ''))
        if mri and lit and m[0].get('rmf', 0) >= rmf_min:
            pts.append((mri, lit))
    if len(pts) < 6:
        return None
    x = np.array([a for a, _ in pts], float); y = np.array([b for _, b in pts], float)
    # Theil-Sen: median of pairwise slopes — robust to outliers (e.g. RI
    # extrapolated below the alkane ladder) without overfitting a subset.
    slopes = [(y[j] - y[i]) / (x[j] - x[i])
              for i in range(len(x)) for j in range(i + 1, len(x))
              if x[j] != x[i]]
    if not slopes:
        return None
    a = float(np.median(slopes))
    b = float(np.median(y - a * x))
    # accept only if the robust fit actually explains most points
    med_resid = float(np.median(np.abs(y - (a * x + b))))
    return (a, b) if med_resid <= 25 else None


def _ri_check(meas_ri, top_name, transform):
    """RI corroboration of the top-1 ID against the per-sample transform.
    OK / suspect      = vs LITERATURE RI (strong);
    OK-pred / suspect-pred = vs CNN-PREDICTED RI, last resort (weak, looser tol);
    noRI = no literature or predicted value; noCal = couldn't calibrate."""
    from config import RI_QC_TOL
    if not meas_ri or not top_name:
        return ''
    if transform is None:
        return 'noCal'
    a, b = transform
    expected = a * meas_ri + b
    lit = _lit_ri(top_name)
    if lit is not None:
        return 'OK' if abs(lit - expected) <= RI_QC_TOL else 'suspect'
    try:
        from config import RI_QC_USE_PREDICTED, RI_QC_PRED_TOL
    except ImportError:
        RI_QC_USE_PREDICTED = False
    if RI_QC_USE_PREDICTED:
        pred = _pred_ri(top_name)
        if pred is not None:
            return 'OK-pred' if abs(pred - expected) <= RI_QC_PRED_TOL else 'suspect-pred'
    return 'noRI'


# ---- Contaminant patterns ----
CONTAMINANT_PATTERNS = [
    'silane', 'siloxane', 'dimethylsilanediol', 'silanediol',
    'sila', 'silyl', 'silox',   # any Si-backbone (trisila…, silyl ethers) = bleed/fiber
    'phthalate', 'phthalic',
    'butylated hydroxytoluene', 'bht',
    'column bleed', 'cyclosiloxane',
    '2,4-di-tert-butylphenol', '2,6-di-tert-butyl',
    'diethyl phthalate', 'dibutyl phthalate', 'diisobutyl phthalate',
    'hexadecanoic acid', 'octadecanoic acid',  # common fatty acids (often contaminants in SPME)
]

# ---- Common food volatiles (auto-confirm candidates) ----
COMMON_FOOD_VOLATILES = [
    'hexanal', 'heptanal', 'octanal', 'nonanal', 'decanal',
    'benzaldehyde', 'phenylacetaldehyde',
    '2-heptenal', '2-octenal', '2-nonenal', '2-decenal',
    '2,4-heptadienal', '2,4-nonadienal', '2,4-decadienal',
    '2-pentylfuran', '2-ethylfuran',
    '1-octen-3-ol', '1-octen-3-one', '1-penten-3-ol',
    '2-heptanone', '2-octanone', '2-nonanone', '2-undecanone',
    'acetophenone', 'beta-ionone', '6,10,14-trimethylpentadecan-2-one',
    'limonene', 'alpha-pinene', 'beta-pinene', 'beta-caryophyllene',
    'linalool', 'alpha-terpineol', 'geraniol',
    '2-methylpyrazine', '2,5-dimethylpyrazine', '2,3,5-trimethylpyrazine',
    '2-ethylpyrazine', '2-ethyl-3,5-dimethylpyrazine',
    'dimethyl disulfide', 'dimethyl trisulfide', 'methional',
    'benzothiazole', 'indole',
    'phenol', 'guaiacol', '4-ethylguaiacol', 'eugenol',
    'hexanoic acid', 'octanoic acid', 'butanoic acid',
    'ethyl hexanoate', 'ethyl octanoate', 'ethyl decanoate',
    'gamma-butyrolactone', 'gamma-hexalactone',
    '2-ethyl-1-hexanol', '1-hexanol', '1-octanol', '1-nonanol',
    '2,2,4-trimethyl-1,3-pentanediol monoisobutyrate',
    '2,2,4-trimethyl-1,3-pentanediol diisobutyrate',
    'furan, 2-pentyl-', 'furan, 2-ethyl-',
    '2,4,6-trimethylpyridine', 'pyridine, 2,4,6-trimethyl-',
    'cyclocitral', 'beta-cyclocitral', 'photocitral',
    'farnesyl acetone', '(e,e)-farnesyl acetone',
    'alpha-isomethylionone', 'isomethylionone',
    '1-hexadecene', '1-octadecene',
    'dodecane', 'tridecane', 'tetradecane', 'pentadecane',
    'hexadecane', 'heptadecane', 'octadecane',
    '3,5-di-tert-butylphenol',
    '2-(2-butoxyethoxy)ethanol', 'ethanol, 1-(2-butoxyethoxy)-',
]


def _is_contaminant(name):
    """Check if compound name matches contaminant patterns."""
    n = name.lower()
    for pat in CONTAMINANT_PATTERNS:
        if pat in n:
            return True
    return False


def _is_common_food(name):
    """Check if compound is a common food volatile."""
    n = name.lower().strip()
    if n in COMMON_FOOD_VOLATILES:
        return True
    # Fuzzy match: if name contains a common food volatile as substring
    for fv in COMMON_FOOD_VOLATILES:
        if len(fv) > 10 and fv in n:
            return True
    return False


def _auto_confirm_status(matches, ri_tolerance=30):
    """
    Determine auto-review status for a peak.

    Returns: ('CONFIRMED', 'REVIEW', 'CONTAMINANT', 'LOW')
    """
    if not matches:
        return ('LOW', 'No match found')

    best = matches[0]
    name = best.get('name', '')
    rmf = best.get('rmf', 0)
    fmf = best.get('fmf', 0)
    ri_diff = best.get('ri_diff')

    # Contaminant check (highest priority)
    if _is_contaminant(name):
        return ('CONTAMINANT', 'Known contaminant / column bleed')

    coelution = best.get('coelution_flag', False)
    ri_ok = ri_diff is not None and ri_diff <= ri_tolerance

    # Auto-confirm: RMF>900 AND FMF>700 (both directions must agree, no co-elution)
    # AND (RI match < 30 OR common food volatile with no RI data)
    if rmf >= 900 and fmf >= 700 and not coelution:
        if ri_ok or (ri_diff is None and _is_common_food(name)):
            return ('CONFIRMED',
                    f'RMF={rmf:.0f} FMF={fmf:.0f} RI={"OK" if ri_ok else "N/A"} food')

    # Low confidence: co-elution or very unbalanced FMF/RMF
    if coelution or (fmf > 0 and rmf / max(fmf, 1) > 1.8):
        return ('LOW', f'Co-elution: RMF/FMF={rmf/max(fmf,1):.1f}')

    # Low confidence: RMF too low
    if rmf < 750 or fmf < 400:
        return ('LOW', f'Low: RMF={rmf:.0f} FMF={fmf:.0f}')

    # Needs review (middle ground)
    return ('REVIEW', f'Review: RMF={rmf:.0f} FMF={fmf:.0f}')


def compile_results(sample_name, integrated_peaks, identification_results,
                    quantification_results, metadata=None):
    """Compile results with top-3 candidates per peak.

    identification_results[i] = [match1, match2, match3, ...]
    """
    rows = []
    # fit this sample's column<->literature RI transform once (self-calibrating)
    _ri_tx = _fit_ri_transform(integrated_peaks, identification_results)
    if _ri_tx:
        print(f"  [RI-QC] self-calibrated transform: litRI={_ri_tx[0]:.3f}*measRI+{_ri_tx[1]:.0f}")

    for i, peak in enumerate(integrated_peaks):
        matches = identification_results[i] if i < len(identification_results) else []
        quant = quantification_results[i] if i < len(quantification_results) else {}

        status, status_reason = _auto_confirm_status(matches)
        top = matches[0] if matches else {}

        # RI QC on the top-1 ID (self-calibrated per-sample transform residual)
        ri_check = _ri_check(peak.get('ri_measured'), top.get('name', ''), _ri_tx)

        # Peak info (shared across all candidates for this peak)
        base = {
            'Sample': sample_name,
            'Peak_No': i + 1,
            'RT_min': round(peak.get('apex_rt', 0), 3),
            'RT_start': round(peak.get('start_rt', 0), 3),
            'RT_end': round(peak.get('end_rt', 0), 3),
            'Area': round(peak.get('area', 0), 1),
            'Height': round(peak.get('height', 0), 1),
            'SN': round(peak.get('sn', 0), 1),
            'Percentage': round(quant.get('percentage', 0), 3) if quant.get('percentage') else '',
            'Status': status,
            'Status_Reason': status_reason,
            'RI_Check': ri_check,   # OK / suspect / noRI (top-1 vs RI transform)
            'User_Confirm': '',  # Empty column for user to fill
        }

        # Check if any known food volatile appears in top-5 (even if not rank 1)
        food_in_top5 = False
        for m in matches[:5]:
            if _is_common_food(m.get('name', '')):
                food_in_top5 = True
                break

        # Top-N candidates as sub-rows
        _mri = peak.get('ri_measured')
        for j, m in enumerate(matches[:TOP_N_CANDIDATES]):
            row = dict(base)
            row['Rank'] = j + 1
            row['Compound_Name'] = m.get('name', 'Unknown')
            row['CAS'] = m.get('cas', '')
            row['FMF'] = m.get('fmf', 0)
            row['RMF'] = m.get('rmf', 0)
            # RI_WAX / RI_Diff use the SAME normalized literature lookup + the
            # per-sample calibration transform as RI_Check, so the three columns
            # never contradict (was: engine's exact-name lookup + raw diff).
            _lw = _lit_ri(m.get('name', ''))
            row['RI_WAX'] = _lw if _lw is not None else ''
            row['RI_DB5'] = m.get('ri_db5', '') if m.get('ri_db5') is not None else ''
            if _lw is not None and _ri_tx and _mri:
                row['RI_Diff'] = int(round(_lw - (_ri_tx[0] * _mri + _ri_tx[1])))
            else:
                row['RI_Diff'] = ''
            row['Source'] = m.get('source', '')
            row['Coelution'] = 'Y' if m.get('coelution_flag') else ''
            row['Food_In_Top5'] = 'Y' if food_in_top5 else ''
            rows.append(row)

        # If no matches, still output one row
        if not matches:
            row = dict(base)
            row['Rank'] = ''
            row['Compound_Name'] = 'Unknown'
            row['CAS'] = ''
            row['FMF'] = ''
            row['RMF'] = ''
            row['RI_Diff'] = ''
            row['RI_Library'] = ''
            row['Coelution'] = ''
            rows.append(row)

    df = pd.DataFrame(rows)

    # ---- 净相对含量% : area-normalized %, but only over non-contaminant peaks
    #      (excludes siloxane/plasticizer/column-bleed so the real components'
    #       share is not diluted by contaminants). Blank for contaminant rows. ----
    if 'Area' in df.columns and 'Status' in df.columns and not df.empty:
        a = pd.to_numeric(df['Area'], errors='coerce').fillna(0)
        clean = df['Status'] != 'CONTAMINANT'
        total_clean = a[(df['Rank'] == 1) & clean].sum()
        if total_clean > 0:
            df['净相对含量%'] = [round(v / total_clean * 100, 3) if c else ''
                                 for v, c in zip(a, clean)]
        else:
            df['净相对含量%'] = ''

    # Sort by Area descending (biggest peaks first, most reliable)
    if 'Area' in df.columns and not df.empty:
        df['_sort_area'] = pd.to_numeric(df['Area'], errors='coerce').fillna(0)
        df['_sort_rank'] = df.groupby('Peak_No')['_sort_area'].transform('max')
        # Stable sort: keep Peak_No ascending within same area
        df = df.sort_values(['_sort_rank', 'Peak_No', 'Rank'],
                            ascending=[False, True, True])
        df = df.drop(columns=['_sort_area', '_sort_rank'])

    return df


def export_to_excel(results_df, output_path, calibration_data=None):
    """Export to color-coded Excel with review workflow."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ---- Build review checklist + summary (analyst workflow: trust green+OK,
    #      review only the uncertain ones, skip contaminants) ----
    r1 = results_df[results_df['Rank'] == 1].copy() if 'Rank' in results_df.columns else results_df.copy()
    ric = r1['RI_Check'] if 'RI_Check' in r1.columns else pd.Series([''] * len(r1), index=r1.index)
    # Review only the YELLOW (needs-review) peaks that are NOT RI-corroborated,
    # plus any RI=suspect (MS says yes but RI contradicts — catch these even if
    # MS-confident). Trust green + RI-OK; skip red(contaminant) & gray(low) in
    # the primary list (gray = low-signal, glance only).
    review_mask = ((((r1['Status'] == 'REVIEW') & (ric != 'OK')) | (ric == 'suspect'))
                   & (r1['Status'] != 'CONTAMINANT'))
    # drop trace peaks below the area-% floor (kept in Results, not actively reviewed)
    try:
        from config import REVIEW_MIN_AREA_PCT as _AREA_FLOOR
    except ImportError:
        _AREA_FLOOR = 0.0
    pct = pd.to_numeric(r1.get('Percentage', pd.Series(index=r1.index, dtype=float)),
                        errors='coerce').fillna(1e9)
    n_trace = int((review_mask & (pct < _AREA_FLOOR)).sum())
    need = r1[review_mask & (pct >= _AREA_FLOOR)]
    rev_cols = [c for c in ['Peak_No', 'RT_min', 'Area', 'Percentage', '净相对含量%',
                            'Area_rel_IS', '相对内标%', 'Status', 'RI_Check',
                            'Compound_Name', 'RMF', 'FMF', 'RI_WAX', 'CAS']
                if c in need.columns]
    review_df = need[rev_cols].sort_values('Area', ascending=False) if not need.empty else need[rev_cols]

    sc = r1['Status'].value_counts()
    rc = ric.value_counts()
    n_total = int(r1['Peak_No'].nunique()) if 'Peak_No' in r1 else len(r1)
    n_contam = int(sc.get('CONTAMINANT', 0))
    n_low = int(sc.get('LOW', 0))
    n_review = int(len(need))
    n_trust = n_total - n_review - n_contam - n_low - n_trace   # green + RI-confirmed
    summary_df = pd.DataFrame({
        '项目': ['峰总数', '可信免审(绿+RI双证据确认)', '★待复核(看"待复核"表)',
                 f'微量峰(<{_AREA_FLOOR}%,免复核)', '污染物(红,可跳过)', '低置信(灰,信号弱,可略看)',
                 '— RI核对OK(文献)', '— RI核对可疑(文献)',
                 '— RI核对OK(预测,弱)', '— RI可疑(预测,弱)', '— 无任何RI可核对'],
        '数量': [n_total, n_trust, n_review, n_trace, n_contam, n_low,
                 int(rc.get('OK', 0)), int(rc.get('suspect', 0)),
                 int(rc.get('OK-pred', 0)), int(rc.get('suspect-pred', 0)),
                 int(rc.get('noRI', 0))],
    })

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='汇总', index=False)
        review_df.to_excel(writer, sheet_name='待复核', index=False)
        results_df.to_excel(writer, sheet_name='Results', index=False)

    # ---- Apply formatting with openpyxl ----
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb['Results']

    # Styles
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hdr_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    body_font = Font(name='Calibri', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Freeze header
    ws.freeze_panes = 'A2'

    # Style header
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Column widths
    widths = {'A': 10, 'B': 6, 'C': 8, 'D': 8, 'E': 8, 'F': 12, 'G': 10, 'H': 6,
              'I': 8, 'J': 10, 'K': 14, 'L': 12, 'M': 30, 'N': 16, 'O': 6, 'P': 6,
              'Q': 8, 'R': 10, 'S': 6, 'T': 12}
    for col_letter, w in widths.items():
        if col_letter in [chr(65+i) for i in range(ws.max_column)]:
            ws.column_dimensions[col_letter].width = w

    # Find status column
    status_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'Status':
            status_col = col
            break

    # Apply row colors based on Status
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=status_col).value if status_col else ''
        fill = gray_fill
        if status == 'CONFIRMED':
            fill = green_fill
        elif status == 'REVIEW':
            fill = yellow_fill
        elif status == 'CONTAMINANT':
            fill = red_fill

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = center_align

        # Bold the Rank=1 row for each peak
        rank_cell = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Rank':
                rank_cell = ws.cell(row=row, column=col)
                break
        if rank_cell and rank_cell.value == 1:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).font = Font(name='Calibri', size=10, bold=True)

    wb.save(output_path)

    # ---- Summary sheet ----
    n_total = results_df['Peak_No'].nunique()
    status_counts = results_df[results_df['Rank'].isin([1, '', None])]['Status'].value_counts()
    n_confirmed = status_counts.get('CONFIRMED', 0)
    n_review = status_counts.get('REVIEW', 0)
    n_contam = status_counts.get('CONTAMINANT', 0)
    n_low = status_counts.get('LOW', 0)

    print(f"  [Export] {output_path}")
    print(f"           {n_total} peaks: "
          f"GREEN={n_confirmed} YELLOW={n_review} RED={n_contam} GRAY={n_low}")
    print(f"           User reviews ~{n_review} peaks (skip {n_confirmed+n_contam+n_low} auto-classified)")
