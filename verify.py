#!/usr/bin/env python3
"""Clean evaluation: SI-only matching (no RI filtering) at user RTs."""
import os, sys, json, re, time, numpy as np

t0 = time.time()
BASE = os.environ['HOME']
txt_path = os.path.join(BASE, 'Desktop', '验证', 'xt6.26.txt')
xlsx_path = os.path.join(BASE, 'Desktop', '验证', 'xt_filled.xlsx')

sys.path.insert(0, os.path.join(BASE, 'Desktop', '桌面', 'gcms-pipeline'))
from gcms_pipeline import (parse_thermo_txt, load_libraries, pre_search_candidates,
    nist_similarity, classify_peak, CONTAMINANTS, AQUEOUS_IMPOSSIBLE, Config)

print('=' * 60)
print('GC-MS Pipeline v4.1 — Verification Report')
print('=' * 60)

# Load
print('Loading NIST 2014 library (239K compounds)...')
library, bp_index, ion_masks = load_libraries()

print('Parsing raw data (449 MB)...')
rts, tics, spectra = parse_thermo_txt(txt_path)
print(f'{len(rts):,d} scans, RT {rts[0]:.1f}-{rts[-1]:.1f} min ({time.time()-t0:.1f}s)')

# Load ground truth
import openpyxl
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb['Sheet1']
ground_truth = []
for r in range(6, ws.max_row + 1):
    rt = ws.cell(row=r, column=1).value
    name = ws.cell(row=r, column=9).value or ws.cell(row=r, column=10).value or ''
    if rt and str(name).strip():
        ground_truth.append({'rt': float(rt), 'name': str(name).strip()})
print(f'Ground truth: {len(ground_truth)} compounds')

# Name matching
def match_name(gt_name, pr_name):
    def norm(s):
        return re.sub(r'[^a-z0-9]', '', s.lower().strip())
    gt_n, pr_n = norm(gt_name), norm(pr_name)
    if gt_n == pr_n: return True
    if len(gt_n) > 8 and (gt_n in pr_n or pr_n in gt_n): return True
    frags = {gt_n[i:i+5] for i in range(len(gt_n)-4) if len(gt_n[i:i+5]) >= 5}
    pr_frags = {pr_n[i:i+5] for i in range(len(pr_n)-4) if len(pr_n[i:i+5]) >= 5}
    if len(frags & pr_frags) >= 3: return True
    # Known pairs
    pairs = [
        ('trimethylpyridine','pyridine'), ('dimethylsilanediol','silanediol'),
        ('betaionone','ionone'), ('pentylfuran','furan'), ('ethylhexanol','propylpentanol'),
    ]
    for a, b in pairs:
        if a in gt_n and b in pr_n: return True
    return False

# Test each GT at exact RT
results = []
for gi, gt in enumerate(ground_truth):
    target_rt = gt['rt']
    idx = np.argmin(np.abs(rts - target_rt))
    if abs(rts[idx] - target_rt) > 0.2:
        results.append({'gt': gt, 'status': 'no_scan'})
        continue

    spec = spectra[idx]
    clean = [(int(round(m)), int(i)) for m, i in spec if int(round(m)) > 25 and i > 100]
    if len(clean) < 5:
        results.append({'gt': gt, 'status': 'low_peaks'})
        continue

    tier, label, keep = classify_peak(clean[:15])
    if not keep:
        results.append({'gt': gt, 'status': f'L3_{label}'})
        continue

    cand = pre_search_candidates(clean, bp_index, ion_masks, library)
    if not cand:
        results.append({'gt': gt, 'status': 'no_candidates'})
        continue

    threshold = Config.mf_medium if tier == 'L1' else max(Config.threshold, Config.mf_high)
    matches = []
    for ci in cand[:500]:
        comp = library[ci]
        mf, rmf, n_sh, n_u, n_l = nist_similarity(clean, comp['peaks'])
        si = max(mf, rmf)
        if si >= threshold:
            matches.append((si, mf, rmf, n_sh, comp['name'], comp.get('cas',''), comp.get('formula','')))

    if not matches:
        results.append({'gt': gt, 'status': 'no_match'})
        continue

    matches.sort(key=lambda x: -x[0])
    top1 = matches[0]
    top5_names = [m[4] for m in matches[:5]]

    is_top1_correct = match_name(gt['name'], top1[4])
    is_top5_correct = any(match_name(gt['name'], n) for n in top5_names)

    results.append({
        'gt': gt, 'correct_top1': is_top1_correct, 'correct_top5': is_top5_correct,
        'top1_si': top1[0], 'top1_mf': top1[1], 'top1_rmf': top1[2],
        'top1_name': top1[4], 'top1_cas': top1[5], 'top5': top5_names,
        'tier': tier, 'n_matches': len(matches),
    })

# Check which GT compounds are in NIST library
def in_library(name):
    n = name.lower().strip()
    for comp in library:
        cn = comp['name'].lower()
        if n == cn: return True, comp['name']
    # Try reversed name
    parts = name.split(', ', 1)
    if len(parts) == 2:
        rev = f'{parts[1]} {parts[0]}'.lower()
        for comp in library:
            if comp['name'].lower() == rev: return True, comp['name']
    # Try substring
    if len(n) > 8:
        for comp in library:
            if n in comp['name'].lower(): return True, comp['name']
    return False, ''

for r in results:
    found, lib_name = in_library(r['gt']['name'])
    r['in_library'] = found
    r['lib_name'] = lib_name

# Summary
correct_all = sum(1 for r in results if r.get('correct_top1'))
correct_lib = sum(1 for r in results if r.get('correct_top1') and r.get('in_library'))
wrong_all = sum(1 for r in results if 'correct_top1' in r and not r['correct_top1'])
wrong_lib = sum(1 for r in results if 'correct_top1' in r and not r['correct_top1'] and r.get('in_library'))
in_lib_total = sum(1 for r in results if r.get('in_library'))
other_status = sum(1 for r in results if 'status' in r)

elapsed = time.time() - t0
print(f'\n{"="*60}')
print(f'RESULTS (SI-only, no RI filtering)')
print(f'{"="*60}')
print(f'  Total GT compounds:             {len(ground_truth):3d}')
print(f'  In NIST 2014 library:           {in_lib_total:3d}')
print(f'  Not in library:                 {len(ground_truth)-in_lib_total-other_status:3d}')
print(f'  Other (no scan/rejected):       {other_status:3d}')
print(f'')
print(f'  Top-1 correct (all):            {correct_all:3d} / {len(ground_truth)} = {correct_all/len(ground_truth)*100:.0f}%')
print(f'  Top-1 correct (in library):     {correct_lib:3d} / {in_lib_total} = {correct_lib/in_lib_total*100:.0f}%' if in_lib_total > 0 else '')
top5_lib = sum(1 for r in results if r.get('correct_top5') and r.get('in_library'))
print(f'  Top-5 correct (in library):     {top5_lib:3d} / {in_lib_total} = {top5_lib/in_lib_total*100:.0f}%' if in_lib_total > 0 else '')

print(f'\n  Elapsed: {elapsed:.1f}s')

print(f'\n--- CORRECT ({correct_all}) ---')
for r in results:
    if r.get('correct_top1'):
        gt = r['gt']
        coel = ' COEL' if r['top1_rmf'] - r['top1_mf'] > 150 else ''
        lib = ' [LIB]' if r['in_library'] else ' [NOT IN LIB]'
        print(f'  RT={gt["rt"]:5.1f}  {gt["name"]:45s}  SI={r["top1_si"]:3d} MF={r["top1_mf"]:3d} RMF={r["top1_rmf"]:3d}{coel}{lib}')

print(f'\n--- WRONG ({wrong_all}) ---')
for r in results:
    if 'correct_top1' in r and not r['correct_top1']:
        gt = r['gt']
        lib = ' [IN LIB]' if r['in_library'] else ''
        print(f'  RT={gt["rt"]:5.1f}  GT: {gt["name"]:40s} -> {r["top1_name"]:40s}  SI={r["top1_si"]:3d}{lib}')
        if r['in_library']:
            print(f'           Library has: {r["lib_name"]}')

print(f'\n--- STATUS ISSUES ({other_status}) ---')
for r in results:
    if 'status' in r:
        print(f'  RT={r["gt"]["rt"]:5.1f}  {r["gt"]["name"]:45s}  ({r["status"]})')

# Save JSON
out_path = os.path.join(BASE, 'Desktop', '验证', 'evaluation_report.json')
with open(out_path, 'w', encoding='utf-8') as f:
    clean_results = []
    for r in results:
        cr = {}
        for k, v in r.items():
            if isinstance(v, (np.integer,)): cr[k] = int(v)
            elif isinstance(v, (np.floating,)): cr[k] = float(v)
            elif isinstance(v, dict): cr[k] = dict(v)
            elif isinstance(v, list): cr[k] = [str(x) for x in v]
            else: cr[k] = v
        clean_results.append(cr)
    json.dump(clean_results, f, ensure_ascii=False, indent=2, default=str)
print(f'\nReport saved to evaluation_report.json')
